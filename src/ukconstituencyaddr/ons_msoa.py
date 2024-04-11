"""
Middle Layers Super Output Areas - small areas 'comprising between 2,000 and 6,000 households and have a usually resident population between 5,000 and 15,000 persons', see https://www.ons.gov.uk/methodology/geography/ukgeographies/censusgeographies/census2021geographies.

Uses data from 2021

https://geoportal.statistics.gov.uk/datasets/ons::local-authority-districts-december-2023-boundaries-uk-bfe-2/about

https://www.ons.gov.uk/census/maps/choropleth/population/age/resident-age-8c/aged-15-to-24-years

For the readble names
https://houseofcommonslibrary.github.io/msoanames/

Filtered by msoa

MSOA = "msoa21"
"""


from copy import copy
import enum
import json
import logging
from typing import List, Optional

import geojson
from matplotlib import cm, pyplot as plt
import openpyxl
import openpyxl.drawing.image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import geopandas as gpd
from sklearn.cluster import KMeans
import sqlalchemy.exc
from sqlalchemy.orm import Session
from shapely.geometry import Point, shape, LineString, MultiLineString
from shapely import ops
import numpy as np
from matplotlib.path import Path
import matplotlib.patheffects as PathEffects

from ukconstituencyaddr import config
from ukconstituencyaddr.db import cacher
from ukconstituencyaddr.db import db_repr_sqlite as db_repr
from ukconstituencyaddr.gis_helpers import combine_gpd_lines
from ukconstituencyaddr.openpyxl_helpers import set_border


class OnsMsoaReadableField(enum.StrEnum):
    """Enum to match fields to headers in the CSV"""

    ID = "msoa21cd"
    NAME = "msoa21nm"
    NAME_W = "msoa21nmw"
    READABLE_NAME = "msoa21hclnm"
    READABLE_NAME_W = "msoa21hclnm"
    LOCAL_AUTHORITY_NAME = "localauthorityname"
    TYPE = "type"


class OnsMsoaCsvParser:
    """Reads ONS Postcode CSV data into the database"""

    def __init__(
        self,
    ) -> None:
        self.csv = config.conf.input.ons_msoa_readble_names_csv
        self.geojson = config.conf.input.ons_msoa_geojson

        for file in [self.geojson, self.csv]:
            if not file.exists():
                raise Exception(f"File not at {file}")

        self.engine = db_repr.get_engine()

        self.logger = logging.getLogger(self.__class__.__name__)

        self.logger.info(f"Using CSV {self.csv}")

    def process_csv(self):
        """Reads the CSV into the database"""
        geojson_modified = cacher.DbCacheInst.check_file_modified(
            cacher.DatafileName.OnsMsoaGeoJson, self.geojson
        )
        readable_name_modified = cacher.DbCacheInst.check_file_modified(
            cacher.DatafileName.OnsMsoaReadableNames, self.csv
        )
        if not geojson_modified and not readable_name_modified:
            self.logger.info("Already parsed CSV and geojson files and placed into db")
            return

        self.logger.info("Parsing ONS MSOA file")

        rows = pd.read_csv(
            self.csv,
            header=0,
            usecols=[
                OnsMsoaReadableField.ID,
                OnsMsoaReadableField.NAME,
                OnsMsoaReadableField.READABLE_NAME,
            ],
        ).set_index(OnsMsoaReadableField.ID)

        # Create empty column for geometry
        rows[db_repr.OnsMsoaColumnsNames.GEOMETRY] = ""
        rows[db_repr.OnsMsoaColumnsNames.GB_OS_EASTING] = 0
        rows[db_repr.OnsMsoaColumnsNames.GB_OS_NORTHING] = 0

        # Add geometry to each entry
        with open(self.geojson) as f:
            msoa_geojson = geojson.load(f)

        msoa_geojson_features = msoa_geojson["features"]
        # Get all geometry objects and add to the dataframe
        for x in msoa_geojson_features:
            msoa_id = x["properties"]["MSOA21CD"]
            easting = x["properties"]["BNG_E"]
            northing = x["properties"]["BNG_N"]
            if msoa_id in rows.index:
                rows.at[msoa_id, db_repr.OnsMsoaColumnsNames.GEOMETRY] = geojson.dumps(
                    x["geometry"]
                )
                rows.at[msoa_id, db_repr.OnsMsoaColumnsNames.GB_OS_EASTING] = easting
                rows.at[msoa_id, db_repr.OnsMsoaColumnsNames.GB_OS_NORTHING] = northing

        rows.rename(
            columns={
                OnsMsoaReadableField.ID: db_repr.OnsMsoaColumnsNames.OID,
                OnsMsoaReadableField.NAME: db_repr.OnsMsoaColumnsNames.NAME,
                OnsMsoaReadableField.READABLE_NAME: db_repr.OnsMsoaColumnsNames.READABLE_NAME,
            },
            inplace=True,
        )
        rows.index.names = [db_repr.OnsMsoaColumnsNames.OID]

        rows.to_sql(
            db_repr.OnsMsoa.__tablename__,
            self.engine,
            if_exists="append",
            index=True,
            chunksize=100000,
        )

        cacher.DbCacheInst.set_file_modified(cacher.DatafileName.OnsMsoaGeoJson, self.geojson)
        cacher.DbCacheInst.set_file_modified(cacher.DatafileName.OnsMsoaReadableNames, self.csv)

        self.logger.info(
            f"Finished parsing ONS MSOA file, wrote {len(rows.index)} items"
        )

    def get_msoa_by_id(self, msoa_id: str) -> Optional[db_repr.OnsMsoa]:
        with Session(self.engine) as session:
            if len(msoa_id) == 0:
                raise ValueError("You must provide a string that isn't empty!")
            else:
                result = (
                    session.query(db_repr.OnsMsoa)
                    .filter(db_repr.OnsMsoa.oid == msoa_id)
                    .one()
                )
                return result

    def get_msoa_by_name(self, msoa_name: str) -> Optional[db_repr.OnsMsoa]:
        with Session(self.engine) as session:
            if len(msoa_name) == 0:
                raise ValueError("You must provide a string that isn't empty!")
            else:
                result = (
                    session.query(db_repr.OnsMsoa)
                    .filter(db_repr.OnsMsoa.name == msoa_name)
                    .one()
                )
                return result

    def get_msoa_by_readable_name(self, msoa_readable_name: str) -> Optional[db_repr.OnsMsoa]:
        with Session(self.engine) as session:
            if len(msoa_readable_name) == 0:
                raise ValueError("You must provide a string that isn't empty!")
            else:
                result = (
                    session.query(db_repr.OnsMsoa)
                    .filter(db_repr.OnsMsoa.readable_name == msoa_readable_name)
                    .one()
                )
                return result

    def clear_all(self):
        """Clears all rows from the ONS MSOA table"""
        with Session(self.engine) as session:
            session.query(db_repr.OnsMsoa).delete()
            session.commit()
            cacher.DbCacheInst.clear_file_modified(cacher.DatafileName.OnsMsoaReadableNames)


class CensusAgeByMsoaFields(enum.StrEnum):
    """Enum to match fields to headers in the CSV"""

    MSOA_ID = "Middle layer Super Output Areas Code"
    MSOA_NAME = "Middle layer Super Output Areas"
    AGE_CATEGORY = "Age (101 categories) Code"
    AGE_CATEGORY_NAME = "Age (101 categories)"
    OBSERVED_COUNT = "Observation"


class CensusAgeByMsoaCsvParser:
    """Reads ONS Postcode CSV data into the database"""

    def __init__(
        self,
    ) -> None:
        self.csv = config.conf.input.census_age_by_msoa_csv
        if not self.csv.exists():
            raise Exception(f"CSV file not at {self.csv}")

        self.csv_name = cacher.DatafileName.CensusAgeByMsoa

        self.engine = db_repr.get_engine()

        self.logger = logging.getLogger(self.__class__.__name__)

        self.logger.info(f"Using CSV {self.csv}")

    def process_csv(self):
        """Reads the CSV into the database"""
        modified = cacher.DbCacheInst.check_and_set_file_modified(
            self.csv_name, self.csv
        )
        if not modified:
            self.logger.info("Already parsed CSV file and placed into db")
            return

        self.logger.info("Parsing Census Age by MSOA file")

        rows = pd.read_csv(
            self.csv,
            dtype={
                CensusAgeByMsoaFields.AGE_CATEGORY: int,
                CensusAgeByMsoaFields.OBSERVED_COUNT: int,
            },
            header=0,
            usecols=[
                CensusAgeByMsoaFields.MSOA_ID,
                CensusAgeByMsoaFields.AGE_CATEGORY,
                CensusAgeByMsoaFields.OBSERVED_COUNT,
            ],
        )

        rows.rename(
            columns={
                CensusAgeByMsoaFields.MSOA_ID: "msoa_id",
                CensusAgeByMsoaFields.AGE_CATEGORY: "age_cat",
                CensusAgeByMsoaFields.OBSERVED_COUNT: "observed_count",
            },
            inplace=True,
        )

        # Create new empty columns
        rows["age_range"] = 0
        rows["percent"] = 0
        rows["sum"] = 0

        # Create a percentage of each age 'category'. Each category is 1 year,
        # e.g. 0 or 16, apart from 100 which means 100+
        rows["percent"] = (
            100
            * rows["observed_count"]
            / rows.groupby("msoa_id")["observed_count"].transform("sum")
        )

        print(rows)

        # Assign a category using bins to each row, e.g. 16 will go in the 15-35 bin
        rows["age_range"] = pd.cut(
            rows["age_cat"],
            [0, 16, 35, 100],
            labels=[
                db_repr.CensusAgeRange.R_0_15,
                db_repr.CensusAgeRange.R_16_35,
                db_repr.CensusAgeRange.R_36_100,
            ],
            include_lowest=True,
        )

        # Create a new dataframe that has observed_count and percent summed for all the age ranges
        new_rows = (
            rows.groupby(["msoa_id", "age_range"])[["observed_count", "percent"]]
            .sum()
            .reset_index()
        )

        # Rename to SQL columns
        new_rows.rename(
            columns={
                "msoa_id": db_repr.CensusAgeByMsoaColumnsNames.MSOA_ID,
                "age_range": db_repr.CensusAgeByMsoaColumnsNames.AGE_RANGE,
                "observed_count": db_repr.CensusAgeByMsoaColumnsNames.OBSERVED_COUNT,
                "percent": db_repr.CensusAgeByMsoaColumnsNames.PERCENT_OF_MSOA,
            },
            inplace=True,
        )

        new_rows.to_sql(
            db_repr.CensusAgeByMsoa.__tablename__,
            self.engine,
            if_exists="append",
            index=False,
            chunksize=100000,
        )

        cacher.DbCacheInst.set_file_modified(self.csv_name, self.csv)

        self.logger.info(
            f"Finished parsing Census Age by MSOA file, wrote {len(rows.index)} items"
        )

    def clear_all(self):
        """Clears all rows from the ONS MSOA table"""
        with Session(self.engine) as session:
            session.query(db_repr.CensusAgeByMsoa).delete()
            session.commit()
            cacher.DbCacheInst.clear_file_modified(self.csv_name)


def get_streets_in_msoa_clustered(msoa_input: str, msoa_parent_dir: Path):
    """
    This is a monolithic function that produces an image, csv and spreadsheet
    for a MSOA

    TODO break down into more readable function
    """
    with Session(db_repr.get_engine()) as session:
        # Try using every kind of name
        found = False
        try:
            msoa = (
                session.query(db_repr.OnsMsoa).filter(db_repr.OnsMsoa.oid == msoa_input).one()
            )
            found = True
        except sqlalchemy.exc.NoResultFound:
            found = False

        if not found:
            try:
                msoa = (
                    session.query(db_repr.OnsMsoa).filter(db_repr.OnsMsoa.name == msoa_input).one()
                )
            except sqlalchemy.exc.NoResultFound:
                found = False
        
        if not found:
            try:
                msoa = (
                    session.query(db_repr.OnsMsoa).filter(db_repr.OnsMsoa.readable_name == msoa_input).one()
                )
            except sqlalchemy.exc.NoResultFound as e:
                raise Exception(f"Unable to find {msoa_input}") from e

        base_filename = f"{msoa.oid} {msoa.readable_name}"

        msoa_dir = msoa_parent_dir / base_filename
        msoa_dir.mkdir(exist_ok=True)

        geometry = json.loads(msoa.geometry)
        msoa_shape = shape(geometry)

        # Read road shape data based on bounds of MSOA
        data = gpd.read_file(
            config.conf.input.os_open_roads_geopackage,
            engine="pyogrio",
            layer="road_link",
            bbox=msoa_shape.bounds,
        )

        # We used square bounds, so now check if every line actually intersects with the MSOA
        rslt_df = data[
            data["geometry"].apply(lambda x: x.intersects(msoa_shape))
        ].reset_index(drop=True)

        # Find the centroids of each line
        rslt_df = rslt_df[["road_classification_number", "name_1", "geometry"]].copy()
        rslt_df = rslt_df.dropna(
            subset=["road_classification_number", "name_1"], how="all"
        ).reset_index(drop=True)

        rslt_df = (
            rslt_df.groupby(["name_1", "road_classification_number"], dropna=False)[
                "geometry"
            ]
            .apply(combine_gpd_lines)
            .reset_index()
        )
        rslt_df["centroid"] = rslt_df["geometry"].apply(lambda x: x.centroid)

        # Convert the list of centroids to tuples of points so that we can use them in scikit
        points = []
        for _, item in rslt_df.iterrows():
            centroid: Point = item["centroid"]
            points.append((centroid.x, centroid.y))

        # Train using rough correct number of 'buckets'
        num_clusters = int(len(points) / 10)
        kmeans = KMeans(n_clusters=num_clusters, random_state=0).fit(points)

        # Create colour mapping for diagram
        color = cm.rainbow(np.linspace(0, 1, num_clusters))
        color_mapping = {x: color[x] for x in range(len(color))}

        fig, ax = plt.subplots()

        # Plot centers
        centers = kmeans.cluster_centers_
        xs = centers[:, 0]
        ys = centers[:, 1]
        cluster_text = range(1, len(xs) + 1)

        # plt.scatter(xs, ys, c="black", s=200, alpha=0.5)
        for x, y, text in zip(xs, ys, cluster_text):
            circle = plt.Circle((x, y), radius=100, color="#9F9F9F")
            ax.add_patch(circle)
            label = ax.annotate(text, xy=(x, y), fontsize=20, color="black", verticalalignment="center", horizontalalignment="center")
            label.set_path_effects([PathEffects.withStroke(linewidth=5, foreground='w')])

        # counter = 0
        # for x, y in zip(xs, ys):
        #     plt.text(x, y, str(counter), , fontsize=12)
        #     counter += 1

        y_kmeans = kmeans.predict(points)

        # Plot on a map
        for idx, item in rslt_df.iterrows():
            this_shape: LineString | MultiLineString = item["geometry"]

            if isinstance(this_shape, MultiLineString):
                lines = this_shape.geoms
            else:
                lines = [this_shape]
            for line in lines:
                x, y = line.coords.xy
                plt.plot(x, y, c=color_mapping[y_kmeans[idx]])

        # Labelling of plot and save to file
        x, y = msoa_shape.exterior.xy
        ax.plot(x, y, c="black")
        ax.set_aspect('equal')
        plt.title(f"{base_filename}")
        plt.axis("off")
        image_file = msoa_dir / f"{base_filename} road clusters.png"
        plt.savefig(image_file, dpi=300)
        plt.clf()

        # Sorting to save to csv
        rslt_df["cluster_number"] = y_kmeans
        rslt_df = rslt_df.drop(["geometry", "centroid"], axis=1)
        rslt_df = rslt_df.sort_values(
            ["cluster_number", "name_1", "road_classification_number"],
            ascending=[True, True, True],
        ).reset_index(drop=True)

        # Cluster index needs changing to 
        rslt_df["cluster_number"] = rslt_df["cluster_number"].apply(lambda x: x + 1)
        rslt_df.rename(
            columns={
                "cluster_number": "Cluster",
                "name_1": "Road name",
                "road_classification_number": "Road classification or number",
            },
            inplace=True,
        )

        # Save to csv
        csv_file = msoa_dir / f"{base_filename} road clusters.csv"
        rslt_df.to_csv(csv_file, index=False)

        # Create the excel workbook
        workbook = openpyxl.Workbook()
        sheet: Worksheet = workbook.active

        sheet.title = msoa.readable_name

        rows = dataframe_to_rows(rslt_df, index=False)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value

        sheet.cell(row=1, column=4).value="Canvassed"
        sheet.cell(row=1, column=5).value="Date"

        header_side = Side(style="medium")
        for x in range(1, 6):
            sheet.cell(row=1, column=x).border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)

        number_of_each_cluster = rslt_df.groupby(["Cluster"]).size()

        current_row_idx = 2
        columns = ["A", "B", "C", "D", "E"]
        for idx, val in number_of_each_cluster.items():
            for col in columns:
                cell_range = f"{col}{current_row_idx}:{col}{current_row_idx+val - 1}"
                set_border(sheet, cell_range, style="thin")
            current_row_idx += val

        sheet["F1"] = f"MSOA ID {base_filename}"
        img = openpyxl.drawing.image.Image(image_file)
        sheet.add_image(img, 'F2')

        workbook_file = msoa_dir / f"{base_filename} road clusters.xlsx"
        workbook.save(filename=workbook_file)
